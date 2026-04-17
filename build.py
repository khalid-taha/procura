“””
build.py — Generate the Order Management workbook from schema.yml

Usage:
python build.py [schema.yml] [output.xlsx]

Defaults:
schema  = schema.yml
output  = OrderManagement.xlsx

Design notes:

- Each table becomes a worksheet with a real Excel Table (ListObject).
- Formula fields use direct cell references (e.g. =E5*F5) so they are
  portable across Excel, LibreOffice, and Google Sheets.
- Lookup fields use INDEX/MATCH against the full column of the referenced
  table — this works even when the referenced table has empty rows.
- Foreign keys and enums are enforced by data validation dropdowns.
- Aggregates on header tables (order totals) use SUMIFS on the detail table.
  “””

import sys
import yaml
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName

NAVY = “0F2B46”
WHITE = “FFFFFF”
GREY = “F8FAFC”
GREEN = “166534”
BLUE = “0000FF”

thin = Side(style=“thin”, color=“CBD5E1”)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

F_TITLE = Font(name=“Arial”, size=16, bold=True, color=NAVY)
F_SUB = Font(name=“Arial”, size=10, color=“64748B”)
F_HEADER = Font(name=“Arial”, size=10, bold=True, color=WHITE)
F_NORMAL = Font(name=“Arial”, size=10, color=“334155”)
F_INPUT = Font(name=“Arial”, size=10, color=BLUE)
F_FORMULA = Font(name=“Arial”, size=10, color=“000000”)
F_LOOKUP = Font(name=“Arial”, size=10, color=GREEN, italic=True)
F_BOLD = Font(name=“Arial”, size=10, bold=True, color=NAVY)

FILL_HEADER = PatternFill(“solid”, fgColor=NAVY)
FILL_KPI = PatternFill(“solid”, fgColor=“F0FDFA”)

ALIGN_LEFT = Alignment(horizontal=“left”, vertical=“center”)
ALIGN_CENTER = Alignment(horizontal=“center”, vertical=“center”)

def col_letter(idx):
return get_column_letter(idx)

def currency_format(symbol):
return f’{symbol}#,##0.00;({symbol}#,##0.00);”-”’

def format_for_type(ftype, fmt_override, symbol):
if fmt_override == “currency” or ftype == “currency”:
return currency_format(symbol)
if ftype == “percent”:
return “0.0%”
if ftype == “date”:
return “DD/MM/YYYY”
if ftype == “number”:
return ‘#,##0.00;(#,##0.00);”-”’
return “@”

def load_schema(path):
with open(path, “r”, encoding=“utf-8”) as f:
return yaml.safe_load(f)

def compute_table_metadata(schema):
default_rows = schema[“workbook”].get(“default_rows”, 500)
data_start = 5  # header at row 4
tables = {}
for t in schema[“tables”]:
fmap = {f[“name”]: idx + 1 for idx, f in enumerate(t[“fields”])}
tables[t[“name”]] = {
“field_map”: fmap,
“data_start”: data_start,
“data_end”: data_start + default_rows - 1,
“fields”: t[“fields”],
“n_cols”: len(t[“fields”]),
}
return tables

def resolve_formula_expr(expr, field_map, row_num):
“”“Replace [@field] tokens with row-specific cell refs (e.g. E5).”””
out = expr
for fname, fidx in field_map.items():
token = f”[@{fname}]”
out = out.replace(token, f”{col_letter(fidx)}{row_num}”)
return out

def resolve_table_refs(expr, table_metadata):
“”“Replace TableName[field] with absolute column ranges.”””
out = expr
for tname, tinfo in table_metadata.items():
for fname, fidx in tinfo[“field_map”].items():
token = f”{tname}[{fname}]”
ref = (f”{tname}!${col_letter(fidx)}${tinfo[‘data_start’]}”
f”:${col_letter(fidx)}${tinfo[‘data_end’]}”)
out = out.replace(token, ref)
return out

def build_vat_rates_sheet(wb, schema):
ws = wb.create_sheet(”_VatRates”)
ws.sheet_state = “hidden”
ws.cell(row=1, column=1, value=“code”).font = F_HEADER
ws.cell(row=1, column=2, value=“rate”).font = F_HEADER
rates = schema.get(“vat_rates”, {})
for i, (code, rate) in enumerate(rates.items(), start=2):
ws.cell(row=i, column=1, value=code)
ws.cell(row=i, column=2, value=rate)
last = len(rates) + 1
wb.defined_names[“VatRates”] = DefinedName(
name=“VatRates”, attr_text=f”_VatRates!$A$2:$B${last}”)

def build_enums_sheet(wb, schema):
ws = wb.create_sheet(”*Enums”)
ws.sheet_state = “hidden”
enums = schema.get(“enums”, {})
for col_idx, (ename, values) in enumerate(enums.items(), start=1):
cl = get_column_letter(col_idx)
ws.cell(row=1, column=col_idx, value=ename).font = F_HEADER
for row_idx, v in enumerate(values, start=2):
ws.cell(row=row_idx, column=col_idx, value=v)
last = len(values) + 1
nm = f”enum*{ename}”
wb.defined_names[nm] = DefinedName(
name=nm, attr_text=f”_Enums!${cl}$2:${cl}${last}”)

def build_table_sheet(wb, table_def, sample_rows, schema, table_metadata):
tname = table_def[“name”]
ws = wb.create_sheet(tname)
if table_def.get(“tab_color”):
ws.sheet_properties.tabColor = table_def[“tab_color”]

```
fields = table_def["fields"]
meta = table_metadata[tname]
field_map = meta["field_map"]
ds, de = meta["data_start"], meta["data_end"]

for i, f in enumerate(fields, 1):
    ws.column_dimensions[get_column_letter(i)].width = f.get("width", 15)

ws.cell(row=1, column=1, value=tname).font = F_TITLE
if table_def.get("description"):
    ws.cell(row=2, column=1, value=table_def["description"]).font = F_SUB

header_row = ds - 1
for col, f in enumerate(fields, 1):
    c = ws.cell(row=header_row, column=col, value=f["name"])
    c.font = F_HEADER
    c.fill = FILL_HEADER
    c.alignment = ALIGN_CENTER
    c.border = BORDER

symbol = schema["workbook"].get("currency_symbol", "£")

for row_offset in range(de - ds + 1):
    row_idx = ds + row_offset
    sample = sample_rows[row_offset] if row_offset < len(sample_rows) else None

    for col, f in enumerate(fields, 1):
        fname = f["name"]
        ftype = f["type"]
        cell = ws.cell(row=row_idx, column=col)

        if ftype == "formula":
            expr = resolve_formula_expr(f["expr"], field_map, row_idx)
            expr = resolve_table_refs(expr, table_metadata)
            # Wrap in IFERROR for clean blanks on empty rows
            inner = expr[1:] if expr.startswith("=") else expr
            cell.value = f'=IFERROR({inner},"")'
            cell.font = F_FORMULA

        elif ftype == "lookup":
            key_field = f["key"]
            key_col_idx = field_map[key_field]
            key_ref = f"{col_letter(key_col_idx)}{row_idx}"
            src_table = f["source"]
            src_meta = table_metadata[src_table]
            src_key_col = src_meta["field_map"][key_field]
            src_tgt_col = src_meta["field_map"][f["field"]]
            key_range = (f"{src_table}!${col_letter(src_key_col)}${src_meta['data_start']}"
                         f":${col_letter(src_key_col)}${src_meta['data_end']}")
            tgt_range = (f"{src_table}!${col_letter(src_tgt_col)}${src_meta['data_start']}"
                         f":${col_letter(src_tgt_col)}${src_meta['data_end']}")
            cell.value = (f'=IFERROR(IF({key_ref}="","",'
                          f'INDEX({tgt_range},MATCH({key_ref},{key_range},0))),"")')
            cell.font = F_LOOKUP

        elif sample and fname in sample and sample[fname] not in ("", None):
            cell.value = sample[fname]
            cell.font = F_INPUT

        elif not sample and "default" in f:
            cell.value = f["default"]
            cell.font = F_INPUT

        else:
            cell.font = F_INPUT

        cell.number_format = format_for_type(ftype, f.get("format"), symbol)
        cell.border = BORDER

# Excel Table
end_col = get_column_letter(len(fields))
table_range = f"A{header_row}:{end_col}{de}"
table = Table(displayName=tname, ref=table_range)
table.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
    showRowStripes=True, showColumnStripes=False)
ws.add_table(table)

# Data validations
for col, f in enumerate(fields, 1):
    cl = get_column_letter(col)
    rng = f"{cl}{ds}:{cl}{de}"
    ftype = f["type"]
    if ftype == "enum":
        dv = DataValidation(type="list", formula1=f'=enum_{f["enum"]}', allow_blank=True)
        dv.error = f"Choose from {f['enum']}"
        dv.errorTitle = "Invalid"
        ws.add_data_validation(dv)
        dv.add(rng)
    elif ftype == "bool":
        dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(rng)
    elif ftype == "fk":
        ref_table, ref_field = f["references"].split(".")
        src_meta = table_metadata[ref_table]
        src_col = src_meta["field_map"][ref_field]
        src_range = (f"={ref_table}!${col_letter(src_col)}${src_meta['data_start']}"
                     f":${col_letter(src_col)}${src_meta['data_end']}")
        dv = DataValidation(type="list", formula1=src_range, allow_blank=True)
        dv.error = f"Must match an existing {ref_table} record"
        dv.errorTitle = "Invalid foreign key"
        ws.add_data_validation(dv)
        dv.add(rng)
```

def build_dashboard(wb, schema, table_metadata):
ws = wb.create_sheet(“Dashboard”)
ws.sheet_properties.tabColor = “DC2626”
ws.column_dimensions[“A”].width = 4
for col in range(2, 8):
ws.column_dimensions[get_column_letter(col)].width = 22

```
ws.cell(row=1, column=2, value=schema["workbook"]["name"]).font = F_TITLE
ws.cell(row=2, column=2, value="Dashboard — KPIs and summary").font = F_SUB

symbol = schema["workbook"].get("currency_symbol", "£")
money = currency_format(symbol)
count = "#,##0"

def rng(tname, fname):
    m = table_metadata[tname]
    c = m["field_map"][fname]
    return (f"{tname}!${col_letter(c)}${m['data_start']}"
            f":${col_letter(c)}${m['data_end']}")

def kpi(row, col, label, formula, is_money=True):
    ws.cell(row=row, column=col, value=label).font = F_SUB
    vc = ws.cell(row=row + 1, column=col, value=formula)
    vc.font = Font(name="Arial", size=18, bold=True, color=NAVY)
    vc.fill = FILL_KPI
    vc.number_format = money if is_money else count
    vc.alignment = ALIGN_LEFT
    vc.border = BORDER

r = 4
ws.cell(row=r, column=2, value="SALES").font = F_BOLD
r += 1
kpi(r, 2, "Total Sales (ex VAT)", f'=SUM({rng("SalesOrders","subtotal")})')
kpi(r, 3, "Sales VAT",            f'=SUM({rng("SalesOrders","vat_total")})')
kpi(r, 4, "Sales (gross)",        f'=SUM({rng("SalesOrders","grand_total")})')
kpi(r, 5, "Received",
    f'=SUMIFS({rng("Payments","amount")},{rng("Payments","direction")},"In")')
kpi(r, 6, "Outstanding (AR)",     f'=SUM({rng("SalesOrders","balance_due")})')

r += 4
ws.cell(row=r, column=2, value="PURCHASES").font = F_BOLD
r += 1
kpi(r, 2, "Total Purchases (ex VAT)", f'=SUM({rng("PurchaseOrders","subtotal")})')
kpi(r, 3, "Purchase VAT",             f'=SUM({rng("PurchaseOrders","vat_total")})')
kpi(r, 4, "Purchases (gross)",        f'=SUM({rng("PurchaseOrders","grand_total")})')
kpi(r, 5, "Paid",
    f'=SUMIFS({rng("Payments","amount")},{rng("Payments","direction")},"Out")')
kpi(r, 6, "Outstanding (AP)",         f'=SUM({rng("PurchaseOrders","balance_due")})')

r += 4
ws.cell(row=r, column=2, value="VAT POSITION").font = F_BOLD
r += 1
kpi(r, 2, "VAT owed to HMRC",
    f'=SUM({rng("SalesOrders","vat_total")})-SUM({rng("PurchaseOrders","vat_total")})')
kpi(r, 3, "Open SOs",
    f'=COUNTIFS({rng("SalesOrders","balance_due")},">0",'
    f'{rng("SalesOrders","status")},"<>Cancelled")', is_money=False)
kpi(r, 4, "Open POs",
    f'=COUNTIFS({rng("PurchaseOrders","balance_due")},">0",'
    f'{rng("PurchaseOrders","status")},"<>Cancelled")', is_money=False)
kpi(r, 5, "Active customers",
    f'=COUNTIFS({rng("Customers","active")},"Yes")', is_money=False)
kpi(r, 6, "Active suppliers",
    f'=COUNTIFS({rng("Suppliers","active")},"Yes")', is_money=False)

r += 4
ws.cell(row=r, column=2, value="SALES ORDERS BY STATUS").font = F_BOLD
r += 1
for col, h in enumerate(["Status", "Count", "Gross Total"], start=2):
    c = ws.cell(row=r, column=col, value=h)
    c.font = F_HEADER
    c.fill = FILL_HEADER
    c.alignment = ALIGN_CENTER
    c.border = BORDER
for i, s in enumerate(schema["enums"]["order_status"]):
    rr = r + 1 + i
    ws.cell(row=rr, column=2, value=s).font = F_NORMAL
    ws.cell(row=rr, column=2).border = BORDER
    ws.cell(row=rr, column=3, value=f'=COUNTIFS({rng("SalesOrders","status")},"{s}")').font = F_FORMULA
    ws.cell(row=rr, column=3).number_format = count
    ws.cell(row=rr, column=3).border = BORDER
    ws.cell(row=rr, column=4,
            value=f'=SUMIFS({rng("SalesOrders","grand_total")},{rng("SalesOrders","status")},"{s}")').font = F_FORMULA
    ws.cell(row=rr, column=4).number_format = money
    ws.cell(row=rr, column=4).border = BORDER

r += len(schema["enums"]["order_status"]) + 3
ws.cell(row=r, column=2, value="PURCHASE ORDERS BY STATUS").font = F_BOLD
r += 1
for col, h in enumerate(["Status", "Count", "Gross Total"], start=2):
    c = ws.cell(row=r, column=col, value=h)
    c.font = F_HEADER
    c.fill = FILL_HEADER
    c.alignment = ALIGN_CENTER
    c.border = BORDER
for i, s in enumerate(schema["enums"]["po_status"]):
    rr = r + 1 + i
    ws.cell(row=rr, column=2, value=s).font = F_NORMAL
    ws.cell(row=rr, column=2).border = BORDER
    ws.cell(row=rr, column=3, value=f'=COUNTIFS({rng("PurchaseOrders","status")},"{s}")').font = F_FORMULA
    ws.cell(row=rr, column=3).number_format = count
    ws.cell(row=rr, column=3).border = BORDER
    ws.cell(row=rr, column=4,
            value=f'=SUMIFS({rng("PurchaseOrders","grand_total")},{rng("PurchaseOrders","status")},"{s}")').font = F_FORMULA
    ws.cell(row=rr, column=4).number_format = money
    ws.cell(row=rr, column=4).border = BORDER
```

def build_stock_levels(wb, schema, table_metadata):
view = next((v for v in schema.get(“views”, []) if v[“name”] == “StockLevels”), None)
if not view:
return
ws = wb.create_sheet(“StockLevels”)
if view.get(“tab_color”):
ws.sheet_properties.tabColor = view[“tab_color”]

```
ws.cell(row=1, column=1, value="Stock Levels").font = F_TITLE
ws.cell(row=2, column=1, value=view.get("description", "")).font = F_SUB

for i, w in enumerate([14, 35, 12, 14, 16], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

for col, h in enumerate(["product_id", "name", "on_hand", "reorder_level", "reorder_status"], 1):
    c = ws.cell(row=4, column=col, value=h)
    c.font = F_HEADER
    c.fill = FILL_HEADER
    c.alignment = ALIGN_CENTER
    c.border = BORDER

prod_meta = table_metadata["Products"]
pds, pde = prod_meta["data_start"], prod_meta["data_end"]

def rng(t, f):
    m = table_metadata[t]
    c = m["field_map"][f]
    return f"{t}!${col_letter(c)}${m['data_start']}:${col_letter(c)}${m['data_end']}"

pid_col = col_letter(prod_meta["field_map"]["product_id"])
pname_col = col_letter(prod_meta["field_map"]["name"])
reorder_col = col_letter(prod_meta["field_map"]["reorder_level"])

for i in range(pde - pds + 1):
    r = 5 + i
    prow = pds + i
    ws.cell(row=r, column=1, value=f'=IF(Products!{pid_col}{prow}="","",Products!{pid_col}{prow})').font = F_LOOKUP
    ws.cell(row=r, column=2, value=f'=IF(Products!{pid_col}{prow}="","",Products!{pname_col}{prow})').font = F_LOOKUP
    ws.cell(row=r, column=3,
            value=f'=IF(A{r}="","",SUMIFS({rng("InventoryMovements","quantity")},{rng("InventoryMovements","product_id")},A{r}))').font = F_FORMULA
    ws.cell(row=r, column=4, value=f'=IF(A{r}="","",Products!{reorder_col}{prow})').font = F_LOOKUP
    ws.cell(row=r, column=5, value=f'=IF(A{r}="","",IF(C{r}<=D{r},"REORDER","OK"))').font = F_FORMULA
    for col in range(1, 6):
        ws.cell(row=r, column=col).border = BORDER
        if col in (3, 4):
            ws.cell(row=r, column=col).number_format = "#,##0"
```

def build_sample_data(schema):
if not schema[“workbook”].get(“sample_data”):
return {}

```
today = date(2026, 4, 1)

customers = [
    {"customer_id": "CUST-0001", "name": "Acme Engineering Ltd",
     "contact_name": "Sarah Mitchell", "email": "orders@acme-eng.co.uk",
     "phone": "020 7946 0101", "billing_address": "12 Industrial Way, Slough SL1 4AB",
     "vat_number": "GB123456789", "credit_limit": 10000, "active": "Yes",
     "created_date": today - timedelta(days=180)},
    {"customer_id": "CUST-0002", "name": "Bright Retail Co",
     "contact_name": "James Patel", "email": "ap@brightretail.com",
     "phone": "0161 496 0823", "billing_address": "Unit 5 Park Trading Est, Manchester M12 4BG",
     "vat_number": "GB987654321", "credit_limit": 5000, "active": "Yes",
     "created_date": today - timedelta(days=120)},
    {"customer_id": "CUST-0003", "name": "Harbour & Sons",
     "contact_name": "Emma Harbour", "email": "emma@harbourandsons.co.uk",
     "phone": "01273 221 094", "billing_address": "8 Marina Court, Brighton BN1 2FG",
     "vat_number": "", "credit_limit": 2500, "active": "Yes",
     "created_date": today - timedelta(days=60)},
    {"customer_id": "CUST-0004", "name": "Northpoint Services",
     "contact_name": "Daniel Okafor", "email": "accounts@northpoint.co.uk",
     "phone": "0113 245 7712", "billing_address": "Regent House, Leeds LS1 5AS",
     "vat_number": "GB445566778", "credit_limit": 7500, "active": "Yes",
     "created_date": today - timedelta(days=30)},
]

suppliers = [
    {"supplier_id": "SUPP-0001", "name": "Precision Components Ltd",
     "contact_name": "Robert Chen", "email": "sales@precisioncomp.co.uk",
     "phone": "01234 567 890", "address": "Factory 3, Bedford MK42 9JJ",
     "vat_number": "GB111222333", "payment_terms_days": 30, "active": "Yes",
     "created_date": today - timedelta(days=200)},
    {"supplier_id": "SUPP-0002", "name": "Global Packaging Supplies",
     "contact_name": "Aisha Khan", "email": "orders@globalpack.com",
     "phone": "0121 789 4561", "address": "Packham Industrial Park, Birmingham B12 8AZ",
     "vat_number": "GB222333444", "payment_terms_days": 45, "active": "Yes",
     "created_date": today - timedelta(days=150)},
    {"supplier_id": "SUPP-0003", "name": "Eastwood Raw Materials",
     "contact_name": "Michael Grant", "email": "sales@eastwoodraw.co.uk",
     "phone": "0191 432 9876", "address": "Dock Road, Newcastle NE1 3DF",
     "vat_number": "GB333444555", "payment_terms_days": 14, "active": "Yes",
     "created_date": today - timedelta(days=90)},
]

products = [
    {"product_id": "PROD-0001", "sku": "WIDGET-A", "name": "Standard Widget (Type A)",
     "description": "10mm steel widget, zinc-plated", "kind": "Both",
     "unit": "ea", "default_sell_price": 12.50, "default_buy_price": 4.80,
     "default_vat_code": "STD", "reorder_level": 50, "active": "Yes"},
    {"product_id": "PROD-0002", "sku": "WIDGET-B", "name": "Premium Widget (Type B)",
     "description": "15mm stainless widget, polished", "kind": "Both",
     "unit": "ea", "default_sell_price": 22.00, "default_buy_price": 9.20,
     "default_vat_code": "STD", "reorder_level": 30, "active": "Yes"},
    {"product_id": "PROD-0003", "sku": "BOX-SM", "name": "Small Packaging Box",
     "description": "200x150x100mm cardboard box", "kind": "Both",
     "unit": "ea", "default_sell_price": 1.20, "default_buy_price": 0.45,
     "default_vat_code": "STD", "reorder_level": 500, "active": "Yes"},
    {"product_id": "PROD-0004", "sku": "SVC-INSTALL", "name": "Installation Service",
     "description": "On-site installation (per hour)", "kind": "Sellable",
     "unit": "hour", "default_sell_price": 85.00, "default_buy_price": 0,
     "default_vat_code": "STD", "reorder_level": 0, "active": "Yes"},
    {"product_id": "PROD-0005", "sku": "RAW-STEEL", "name": "Raw Steel Sheet",
     "description": "1.2mm x 1m x 2m sheet", "kind": "Purchasable",
     "unit": "sheet", "default_sell_price": 0, "default_buy_price": 48.00,
     "default_vat_code": "STD", "reorder_level": 20, "active": "Yes"},
    {"product_id": "PROD-0006", "sku": "CONSUMABLE", "name": "Workshop Consumables Pack",
     "description": "Gloves, cleaning fluids, fasteners", "kind": "Purchasable",
     "unit": "pack", "default_sell_price": 0, "default_buy_price": 35.00,
     "default_vat_code": "STD", "reorder_level": 5, "active": "Yes"},
]

sales_orders = [
    {"order_id": "SO-0001", "order_date": today - timedelta(days=25),
     "customer_id": "CUST-0001", "status": "Paid",
     "required_date": today - timedelta(days=15),
     "shipped_date": today - timedelta(days=18),
     "po_reference": "ACME-PO-8821", "notes": ""},
    {"order_id": "SO-0002", "order_date": today - timedelta(days=18),
     "customer_id": "CUST-0002", "status": "Invoiced",
     "required_date": today - timedelta(days=8),
     "shipped_date": today - timedelta(days=10),
     "po_reference": "BR-2026-0418", "notes": "Deliver to goods-in"},
    {"order_id": "SO-0003", "order_date": today - timedelta(days=10),
     "customer_id": "CUST-0003", "status": "Shipped",
     "required_date": today - timedelta(days=2),
     "shipped_date": today - timedelta(days=3),
     "po_reference": "", "notes": "Phone order"},
    {"order_id": "SO-0004", "order_date": today - timedelta(days=3),
     "customer_id": "CUST-0004", "status": "Confirmed",
     "required_date": today + timedelta(days=7),
     "shipped_date": "", "po_reference": "NP-2026-0088", "notes": ""},
    {"order_id": "SO-0005", "order_date": today,
     "customer_id": "CUST-0001", "status": "Draft",
     "required_date": today + timedelta(days=14),
     "shipped_date": "", "po_reference": "", "notes": "Awaiting confirmation"},
]

sales_lines = [
    {"line_id": "SOL-0001", "order_id": "SO-0001", "product_id": "PROD-0001",
     "quantity": 100, "unit_price": 12.50, "discount_pct": 0, "vat_code": "STD"},
    {"line_id": "SOL-0002", "order_id": "SO-0001", "product_id": "PROD-0003",
     "quantity": 10, "unit_price": 1.20, "discount_pct": 0, "vat_code": "STD"},
    {"line_id": "SOL-0003", "order_id": "SO-0002", "product_id": "PROD-0002",
     "quantity": 50, "unit_price": 22.00, "discount_pct": 0.05, "vat_code": "STD"},
    {"line_id": "SOL-0004", "order_id": "SO-0003", "product_id": "PROD-0001",
     "quantity": 24, "unit_price": 12.50, "discount_pct": 0, "vat_code": "STD"},
    {"line_id": "SOL-0005", "order_id": "SO-0003", "product_id": "PROD-0004",
     "quantity": 4, "unit_price": 85.00, "discount_pct": 0, "vat_code": "STD"},
    {"line_id": "SOL-0006", "order_id": "SO-0004", "product_id": "PROD-0002",
     "quantity": 30, "unit_price": 22.00, "discount_pct": 0.10, "vat_code": "STD"},
    {"line_id": "SOL-0007", "order_id": "SO-0004", "product_id": "PROD-0003",
     "quantity": 30, "unit_price": 1.20, "discount_pct": 0, "vat_code": "STD"},
    {"line_id": "SOL-0008", "order_id": "SO-0005", "product_id": "PROD-0001",
     "quantity": 200, "unit_price": 12.00, "discount_pct": 0, "vat_code": "STD"},
]

purchase_orders = [
    {"po_id": "PO-0001", "po_date": today - timedelta(days=60),
     "supplier_id": "SUPP-0001", "status": "Paid",
     "expected_date": today - timedelta(days=45),
     "received_date": today - timedelta(days=47),
     "supplier_invoice_ref": "PC-INV-44221", "notes": ""},
    {"po_id": "PO-0002", "po_date": today - timedelta(days=35),
     "supplier_id": "SUPP-0002", "status": "Invoiced",
     "expected_date": today - timedelta(days=20),
     "received_date": today - timedelta(days=22),
     "supplier_invoice_ref": "GP-0119-26", "notes": ""},
    {"po_id": "PO-0003", "po_date": today - timedelta(days=12),
     "supplier_id": "SUPP-0003", "status": "Received",
     "expected_date": today - timedelta(days=5),
     "received_date": today - timedelta(days=4),
     "supplier_invoice_ref": "", "notes": "Invoice awaited"},
    {"po_id": "PO-0004", "po_date": today - timedelta(days=2),
     "supplier_id": "SUPP-0001", "status": "Sent",
     "expected_date": today + timedelta(days=12),
     "received_date": "", "supplier_invoice_ref": "", "notes": ""},
]

po_lines = [
    {"line_id": "POL-0001", "po_id": "PO-0001", "product_id": "PROD-0001",
     "quantity": 500, "unit_price": 4.80, "vat_code": "STD", "quantity_received": 500},
    {"line_id": "POL-0002", "po_id": "PO-0001", "product_id": "PROD-0002",
     "quantity": 200, "unit_price": 9.20, "vat_code": "STD", "quantity_received": 200},
    {"line_id": "POL-0003", "po_id": "PO-0002", "product_id": "PROD-0003",
     "quantity": 2000, "unit_price": 0.45, "vat_code": "STD", "quantity_received": 2000},
    {"line_id": "POL-0004", "po_id": "PO-0003", "product_id": "PROD-0005",
     "quantity": 40, "unit_price": 48.00, "vat_code": "STD", "quantity_received": 40},
    {"line_id": "POL-0005", "po_id": "PO-0003", "product_id": "PROD-0006",
     "quantity": 10, "unit_price": 35.00, "vat_code": "STD", "quantity_received": 10},
    {"line_id": "POL-0006", "po_id": "PO-0004", "product_id": "PROD-0001",
     "quantity": 300, "unit_price": 4.80, "vat_code": "STD", "quantity_received": 0},
]

payments = [
    {"payment_id": "PAY-0001", "payment_date": today - timedelta(days=15),
     "direction": "In", "reference_id": "SO-0001", "amount": 1264.80,
     "method": "Bank Transfer", "reference": "ACME-REMIT-0918", "notes": ""},
    {"payment_id": "PAY-0002", "payment_date": today - timedelta(days=6),
     "direction": "In", "reference_id": "SO-0002", "amount": 800.00,
     "method": "Bank Transfer", "reference": "BR-PART-001", "notes": "Part payment"},
    {"payment_id": "PAY-0003", "payment_date": today - timedelta(days=40),
     "direction": "Out", "reference_id": "PO-0001", "amount": 5088.00,
     "method": "Bank Transfer", "reference": "BACS-26-0219", "notes": ""},
    {"payment_id": "PAY-0004", "payment_date": today - timedelta(days=8),
     "direction": "Out", "reference_id": "PO-0002", "amount": 1080.00,
     "method": "Bank Transfer", "reference": "BACS-26-0404", "notes": ""},
]

inventory = [
    {"movement_id": "MOV-0001", "movement_date": today - timedelta(days=47),
     "product_id": "PROD-0001", "quantity": 500, "movement_type": "Receipt",
     "reference_id": "PO-0001", "notes": ""},
    {"movement_id": "MOV-0002", "movement_date": today - timedelta(days=47),
     "product_id": "PROD-0002", "quantity": 200, "movement_type": "Receipt",
     "reference_id": "PO-0001", "notes": ""},
    {"movement_id": "MOV-0003", "movement_date": today - timedelta(days=22),
     "product_id": "PROD-0003", "quantity": 2000, "movement_type": "Receipt",
     "reference_id": "PO-0002", "notes": ""},
    {"movement_id": "MOV-0004", "movement_date": today - timedelta(days=4),
     "product_id": "PROD-0005", "quantity": 40, "movement_type": "Receipt",
     "reference_id": "PO-0003", "notes": ""},
    {"movement_id": "MOV-0005", "movement_date": today - timedelta(days=4),
     "product_id": "PROD-0006", "quantity": 10, "movement_type": "Receipt",
     "reference_id": "PO-0003", "notes": ""},
    {"movement_id": "MOV-0006", "movement_date": today - timedelta(days=18),
     "product_id": "PROD-0001", "quantity": -100, "movement_type": "Dispatch",
     "reference_id": "SO-0001", "notes": ""},
    {"movement_id": "MOV-0007", "movement_date": today - timedelta(days=18),
     "product_id": "PROD-0003", "quantity": -10, "movement_type": "Dispatch",
     "reference_id": "SO-0001", "notes": ""},
    {"movement_id": "MOV-0008", "movement_date": today - timedelta(days=10),
     "product_id": "PROD-0002", "quantity": -50, "movement_type": "Dispatch",
     "reference_id": "SO-0002", "notes": ""},
    {"movement_id": "MOV-0009", "movement_date": today - timedelta(days=3),
     "product_id": "PROD-0001", "quantity": -24, "movement_type": "Dispatch",
     "reference_id": "SO-0003", "notes": ""},
]

return {
    "Customers": customers,
    "Suppliers": suppliers,
    "Products": products,
    "SalesOrders": sales_orders,
    "SalesOrderLines": sales_lines,
    "PurchaseOrders": purchase_orders,
    "PurchaseOrderLines": po_lines,
    "Payments": payments,
    "InventoryMovements": inventory,
}
```

def main():
schema_path = sys.argv[1] if len(sys.argv) > 1 else “schema.yml”
output_path = sys.argv[2] if len(sys.argv) > 2 else “OrderManagement.xlsx”

```
print(f"Loading schema: {schema_path}")
schema = load_schema(schema_path)

wb = Workbook()
wb.remove(wb.active)

table_metadata = compute_table_metadata(schema)

build_vat_rates_sheet(wb, schema)
build_enums_sheet(wb, schema)

samples = build_sample_data(schema)

for t in schema["tables"]:
    n = len(samples.get(t["name"], []))
    print(f"  Building table: {t['name']}  ({n} sample rows)")
    build_table_sheet(wb, t, samples.get(t["name"], []), schema, table_metadata)

print("  Building Dashboard")
build_dashboard(wb, schema, table_metadata)

print("  Building views")
build_stock_levels(wb, schema, table_metadata)

table_names = [t["name"] for t in schema["tables"]]
view_names = [v["name"] for v in schema.get("views", []) if v["name"] != "Dashboard"]
desired_order = (
    ["Dashboard"] + table_names + view_names + ["_VatRates", "_Enums"]
)
wb._sheets = [wb[name] for name in desired_order if name in wb.sheetnames]

wb.save(output_path)
print(f"Saved: {output_path}")
```

if **name** == “**main**”:
main()
